#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
import re
import shutil
import sys
import unicodedata
import urllib.request
import zipfile
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from html import unescape
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl não está instalado. Instale requirements.txt ou use o Python empacotado do Codex."
    ) from exc


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "saida_bling"
CLASSIF_URL = (
    "https://portalunico.siscomex.gov.br/classif/api/publico/nomenclatura/download/json"
)

EXPECTED_PRODUCT_COUNT = 78
EXPECTED_TOTAL_QTY = Decimal("241")
EXPECTED_TOTAL_COST = Decimal("11884.94")
EXPECTED_TOTAL_SALE = Decimal("56582.93")
MONEY_TOLERANCE = Decimal("0.05")

NAMESPACES = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}

MAIN_COLUMNS = [
    "linha_ods",
    "status_cadastro",
    "alertas",
    "nome_original",
    "nome_bling",
    "sku",
    "descricao_curta",
    "marca",
    "modelo_detectado",
    "tipo_peca",
    "categoria_path",
    "categoria_bling_id",
    "quantidade",
    "custo_unitario",
    "preco_venda",
    "margem_pct",
    "valor_estoque_venda",
    "unidade",
    "formato_produto",
    "situacao_sugerida",
    "peso_liquido",
    "peso_bruto",
    "largura",
    "altura",
    "profundidade",
    "gtin",
    "ncm",
    "ncm_status",
    "ncm_fonte",
    "ncm_descricao_oficial",
    "url_imagem",
    "bling_id",
    "status_importacao",
    "erro_importacao",
]

EXPECTED_DISTRIBUTION = {
    "DJI > Peças Originais DJI > Frames e Carcaças": 12,
    "DJI > Peças Originais DJI > Dobradiças Eixos e Acabamentos": 17,
    "DJI > Peças Originais DJI > Braços": 16,
    "DJI > Peças Originais DJI > Gimbals PTZ e Cabos": 7,
    "DJI > Peças Originais DJI > Placas ESC e Controladoras": 7,
    "DJI > Peças Originais DJI > Câmeras e CMOS": 4,
    "DJI > Peças Originais DJI > Sensores IMU e GPS": 4,
    "DJI > Peças Originais DJI > Hélices e Rotores": 4,
    "DJI > Peças Originais DJI > Controles Remotos": 2,
    "DJI > Peças Originais DJI > Carregadores e Hubs": 2,
    "DJI > Peças Originais DJI > Películas e Proteções": 1,
    "DJI > Peças Originais DJI > Baterias e Tampas": 1,
    "DJI > Drones Completos": 1,
}

CATEGORY_ALIASES = {
    "DJI > Peças Originais DJI > Controles Remotos": [
        "DJI > Acessórios DJI > Controles Remotos"
    ],
    "DJI > Peças Originais DJI > Carregadores e Hubs": [
        "DJI > Acessórios DJI > Carregadores e Hubs"
    ],
    "DJI > Peças Originais DJI > Películas e Proteções": [
        "DJI > Acessórios DJI > Películas e Proteções"
    ],
    "DJI > Peças Originais DJI > Dobradiças Eixos e Acabamentos": [
        "DJI > Peças Originais DJI > Dobradiças, Eixos e Acabamentos"
    ],
    "DJI > Peças Originais DJI > Gimbals PTZ e Cabos": [
        "DJI > Peças Originais DJI > Gimbals, PTZ e Cabos"
    ],
    "DJI > Peças Originais DJI > Placas ESC e Controladoras": [
        "DJI > Peças Originais DJI > Placas, ESC e Controladoras"
    ],
    "DJI > Peças Originais DJI > Sensores IMU e GPS": [
        "DJI > Peças Originais DJI > Sensores, IMU e GPS"
    ],
}

MODEL_PATTERNS = [
    ("Mavic 3 Classic", r"\bmavic\s*3\s*classic\b"),
    ("Mavic 3 Pro", r"\bmavic\s*3\s*pro\b"),
    ("Mavic 3", r"\bmavic\s*3\b"),
    ("Mavic 2 Zoom", r"\bmavic\s*(?:pro\s*)?2\s*(?:zoom|zom)\b"),
    ("Mavic 2 Pro", r"\bmavic\s*2\s*pro\b"),
    ("Mavic 2", r"\bmavic\s*2\b"),
    ("Mini 5 Pro", r"\bmini\s*5\s*pro\b"),
    ("Mini 4 Pro", r"\bmini\s*4\s*pro\b"),
    ("Mini 4", r"\bmini\s*4\b"),
    ("Mini 3 Pro", r"\bmini\s*3\s*pro\b"),
    ("Mini 3", r"\bmini\s*3\b"),
    ("Mini 2 SE", r"\bmini\s*2\s*se\b"),
    ("Mini 2", r"\bmini\s*2\b"),
    ("Air 3S", r"\bair\s*3s\b"),
    ("Air 3", r"\bair\s*3\b"),
    ("Air 2S", r"\bair\s*2s\b"),
    ("Air 2", r"\bair\s*2\b"),
    ("Avata 2", r"\bavata\s*2\b"),
    ("Avata", r"\bavata\b"),
    ("Neo 2", r"\bneo\s*2\b"),
    ("Neo", r"\bneo\b"),
    ("Flip", r"\bflip\b"),
    ("FPV", r"\bfpv\b"),
    ("Inspire 3", r"\binspire\s*3\b"),
    ("Inspire 2", r"\binspire\s*2\b"),
    ("Matrice", r"\bmatrice\b"),
    ("Agras", r"\bagras\b"),
]

MODEL_SKU = {
    "Mavic 3 Classic": "MAVIC3CLASSIC",
    "Mavic 3 Pro": "MAVIC3PRO",
    "Mavic 3": "MAVIC3",
    "Mavic 2 Zoom": "MAVIC2ZOOM",
    "Mavic 2 Pro": "MAVIC2PRO",
    "Mavic 2": "MAVIC2",
    "Mini 5 Pro": "MINI5PRO",
    "Mini 4 Pro": "MINI4PRO",
    "Mini 4": "MINI4",
    "Mini 3 Pro": "MINI3PRO",
    "Mini 3": "MINI3",
    "Mini 2 SE": "MINI2SE",
    "Mini 2": "MINI2",
    "Air 3S": "AIR3S",
    "Air 3": "AIR3",
    "Air 2S": "AIR2S",
    "Air 2": "AIR2",
    "Avata 2": "AVATA2",
    "Avata": "AVATA",
    "Neo 2": "NEO2",
    "Neo": "NEO",
    "Flip": "FLIP",
    "FPV": "FPV",
    "Inspire 3": "INSPIRE3",
    "Inspire 2": "INSPIRE2",
    "Matrice": "MATRICE",
    "Agras": "AGRAS",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def normalize_key(value: str) -> str:
    text = strip_accents(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def strip_accents(value: str) -> str:
    return "".join(
        c
        for c in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(c)
    )


def sku_part(value: str, fallback: str = "GERAL") -> str:
    text = strip_accents(value).upper()
    text = re.sub(r"[^A-Z0-9]+", "-", text).strip("-")
    return text or fallback


def parse_decimal(value) -> Decimal | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", "").replace("R$", "").strip()
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except Exception:
        return None


def read_ods_rows(path: Path) -> list[dict]:
    if not path.exists():
        raise SystemExit(f"Arquivo não encontrado: {path}")

    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("content.xml"))

    raw_rows: list[list[str]] = []
    for table in root.findall(".//table:table", NAMESPACES):
        for tr in table.findall("table:table-row", NAMESPACES):
            row: list[str] = []
            for cell in tr.findall("table:table-cell", NAMESPACES):
                repeat = int(
                    cell.attrib.get(
                        f"{{{NAMESPACES['table']}}}number-columns-repeated", "1"
                    )
                )
                value = cell.attrib.get(f"{{{NAMESPACES['office']}}}value")
                if value is None:
                    texts = [
                        "".join(p.itertext())
                        for p in cell.findall(".//text:p", NAMESPACES)
                    ]
                    value = "\n".join(t for t in texts if t).strip()
                row.extend([str(value)] * min(repeat, 10))
            if any(str(c).strip() for c in row):
                raw_rows.append(row[:4])

    products: list[dict] = []
    for row_idx, row in enumerate(raw_rows[1:], start=2):
        if not row or not str(row[0]).strip():
            continue
        products.append(
            {
                "linha_ods": row_idx,
                "nome_original": str(row[0]).strip(),
                "quantidade": parse_decimal(row[1] if len(row) > 1 else None),
                "custo_unitario": parse_decimal(row[2] if len(row) > 2 else None),
                "preco_venda": parse_decimal(row[3] if len(row) > 3 else None),
            }
        )
    return products


def download_ncm_table() -> tuple[dict[str, dict], str]:
    request = urllib.request.Request(
        CLASSIF_URL,
        headers={
            "User-Agent": "ZalenShopBrasilDrones/1.0",
            "Accept": "application/json",
        },
    )
    downloaded_at = now_iso()
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = json.loads(response.read().decode("utf-8"))

    if isinstance(payload, dict):
        items = payload.get("Nomenclaturas") or payload.get("data") or payload.get("items")
    else:
        items = payload

    if not isinstance(items, list):
        raise SystemExit("Formato inesperado no JSON oficial NCM/Classif.")

    index: dict[str, dict] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        code = str(item.get("Codigo") or item.get("codigo") or "").strip()
        if not code:
            continue
        digits = re.sub(r"\D", "", code)
        description = unescape(str(item.get("Descricao") or item.get("descricao") or ""))
        description = re.sub(r"<[^>]+>", "", description)
        clean = {
            "codigo": code,
            "codigo_digitos": digits,
            "descricao": re.sub(r"\s+", " ", description).strip(),
            "data_inicio": item.get("Data_Inicio") or item.get("dataInicio"),
            "data_fim": item.get("Data_Fim") or item.get("dataFim"),
        }
        index[code] = clean
        if digits:
            index[digits] = clean
    return index, downloaded_at


def resolve_ods_path() -> Path:
    explicit = os.getenv("BLING_ODS_PATH")
    if explicit:
        candidate = Path(explicit).expanduser()
        if candidate.exists():
            return candidate
        raise SystemExit(f"BLING_ODS_PATH apontou para um arquivo inexistente: {candidate}")

    if len(sys.argv) > 1:
        candidate = Path(sys.argv[1]).expanduser()
        if candidate.exists():
            return candidate
        raise SystemExit(f"Arquivo ODS informado na linha de comando não existe: {candidate}")

    candidates = [
        Path.cwd() / "CADASTPRODUTOS.ods",
        ROOT / "CADASTPRODUTOS.ods",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    raise SystemExit(
        "Arquivo ODS não encontrado. Use BLING_ODS_PATH=/caminho/CADASTPRODUTOS.ods "
        "ou passe o caminho como primeiro argumento."
    )


def resolve_category_map_path() -> Path | None:
    explicit = os.getenv("BLING_CATEGORY_MAP_PATH")
    if explicit:
        candidate = Path(explicit).expanduser()
        if candidate.exists():
            return candidate
        raise SystemExit(
            f"BLING_CATEGORY_MAP_PATH apontou para um arquivo inexistente: {candidate}"
        )

    candidates = [
        OUTPUT_DIR / "category-map.json",
        ROOT / "scripts" / "bling" / "logs" / "latest-category-map.json",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def load_category_map(category_map_path: Path | None) -> dict[str, int | None]:
    if category_map_path and category_map_path.exists():
        document = json.loads(category_map_path.read_text("utf-8"))
        source = document.get("categoryMap", document)
    else:
        source = {}

    by_normalized = {normalize_key(k): v for k, v in source.items()}
    result: dict[str, int | None] = {}

    for path in EXPECTED_DISTRIBUTION:
        found = by_normalized.get(normalize_key(path))
        if found is None:
            for alias in CATEGORY_ALIASES.get(path, []):
                found = by_normalized.get(normalize_key(alias))
                if found is not None:
                    break
        result[path] = int(found) if found is not None else None

    return result


def detect_model(name: str) -> tuple[str, list[str]]:
    normalized = normalize_key(name)
    alerts: list[str] = []
    for model, pattern in MODEL_PATTERNS:
        if re.search(pattern, normalized):
            if model == "Mini 5 Pro":
                alerts.append("modelo fora do mapeamento esperado: Mini 5 Pro")
            if model == "Mini 4":
                alerts.append("revisar modelo: Mini 4 ou Mini 4 Pro")
            if model == "Mini 3":
                alerts.append("revisar modelo: Mini 3 ou Mini 3 Pro")
            return model, alerts
    alerts.append("modelo não detectado")
    return "", alerts


def infer_type_position(name: str) -> tuple[str, str, str, list[str]]:
    n = normalize_key(name)
    alerts: list[str] = []

    if any(k in n for k in ["pelicula", "protecao", "protector", "guard"]):
        return "Película e Proteção", "PELICULA", "", alerts
    if any(k in n for k in ["carregador", "charger", "hub"]):
        return "Hub Carregador", "HUB", "", alerts
    if any(k in n for k in ["controle", "remote", " rc ", "radio", "controller"]):
        return "Controle Remoto", "CONTROLE", "", alerts
    if "dji neo 2" in n:
        return "Drone", "DRONE", "", alerts
    if "kit core" in n and "cmos" in n:
        return "Kit Core com CMOS", "KIT-CORE-CMOS", "", alerts
    if "cabo ptz" in n or "cabo gimbal" in n or "flat gimbal" in n:
        return "Cabo PTZ", "CABO-PTZ", "", alerts
    if "gimbal" in n or "ptz" in n:
        return "Gimbal com PTZ", "GIMBAL-PTZ", "", alerts
    if "cmos" in n or "camera" in n or "camera" in strip_accents(n):
        return "CMOS", "CMOS", "", alerts
    if "gps" in n or "imu" in n or "sensor" in n or "vision" in n:
        return "GPS IMU", "GPS-IMU", "", alerts
    if "placa esc" in n or re.search(r"\besc\b", n):
        return "Placa ESC", "PLACA-ESC", "", alerts
    if "placa core" in n or "core" in n or "main" in n or "controladora" in n:
        return "Placa Core", "PLACA-CORE", "", alerts
    if "helice" in n or "rotor" in n or "propeller" in n:
        return "Hélice", "HELICE", "", alerts
    if "tampa da bateria" in n or "tampa bateria" in n:
        return "Tampa da Bateria", "TAMPA-BAT", "", alerts
    if "braco" in n or "arm" in n:
        return "Braço", "BRACO", detect_position(n), alerts
    if "dobradica" in n:
        return "Dobradiça", "DOBRADICA", detect_position(n), alerts
    if "eixo" in n:
        return "Eixo Dobradiça", "EIXO-DOB", detect_position(n), alerts
    if "acabamento" in n:
        return "Acabamento", "ACABAMENTO", detect_position(n), alerts
    if "shell" in n or "carcaca" in n or "frame" in n or "cover" in n:
        return "Shell", "SHELL", detect_position(n), alerts

    alerts.append("nome abreviado demais")
    return "Peça", "PECA", "", alerts


def detect_position(normalized_name: str) -> str:
    parts = []
    if "dianteir" in normalized_name or re.search(r"\bdia\b", normalized_name):
        parts.append("Dianteiro")
    if "traseir" in normalized_name or re.search(r"\btra\b", normalized_name):
        parts.append("Traseiro")
    if "direit" in normalized_name or re.search(r"\bdir\b", normalized_name):
        parts.append("Direito")
    if "esquerd" in normalized_name or re.search(r"\besq\b", normalized_name):
        parts.append("Esquerdo")
    if "inferior" in normalized_name or "infeior" in normalized_name:
        parts.append("Inferior")
    if "superior" in normalized_name:
        parts.append("Superior")
    if "meio" in normalized_name:
        parts.append("Meio")
    return " ".join(parts)


def infer_category(name: str) -> str:
    n = normalize_key(name)
    if "dji neo 2" in n:
        return "DJI > Drones Completos"
    if any(k in n for k in ["pelicula", "protecao", "protector", "guard"]):
        return "DJI > Peças Originais DJI > Películas e Proteções"
    if any(k in n for k in ["carregador", "charger", "hub"]):
        return "DJI > Peças Originais DJI > Carregadores e Hubs"
    if any(k in n for k in ["controle", "remote", " rc ", "radio", "controller"]):
        return "DJI > Peças Originais DJI > Controles Remotos"
    if any(k in n for k in ["shell", "carcaca", "tampa", "cover", "frame"]):
        if "tampa da bateria" in n or "tampa bateria" in n:
            return "DJI > Peças Originais DJI > Baterias e Tampas"
        return "DJI > Peças Originais DJI > Frames e Carcaças"
    if any(k in n for k in ["dobradica", "eixo", "haste", "acabamento", "trava"]):
        return "DJI > Peças Originais DJI > Dobradiças Eixos e Acabamentos"
    if "braco" in n or "arm" in n:
        return "DJI > Peças Originais DJI > Braços"
    if any(k in n for k in ["gimbal", "ptz", "flat gimbal", "cabo gimbal", "cabo ptz"]):
        return "DJI > Peças Originais DJI > Gimbals PTZ e Cabos"
    if "camera" in n or "cmos" in n:
        return "DJI > Peças Originais DJI > Câmeras e CMOS"
    if any(k in n for k in ["placa", "board", "esc", "core", "main", "controladora"]):
        return "DJI > Peças Originais DJI > Placas ESC e Controladoras"
    if any(k in n for k in ["sensor", "imu", "gps", "vision"]):
        return "DJI > Peças Originais DJI > Sensores IMU e GPS"
    if any(k in n for k in ["helice", "rotor", "propeller"]):
        return "DJI > Peças Originais DJI > Hélices e Rotores"
    return "DJI > Outros a Classificar"


def build_name(type_label: str, position: str, model: str, original: str) -> str:
    if type_label == "Drone" and model:
        return f"Drone DJI {model}"
    if type_label == "Película e Proteção":
        suffix = f" DJI {model}" if model else " DJI"
        if "controle" in normalize_key(original):
            return f"Película para Controle{suffix}"
        return f"Película de Proteção{suffix}"
    if type_label == "Hub Carregador":
        return f"Hub Carregador DJI {model}".strip()
    if type_label == "Controle Remoto":
        return f"Controle Remoto DJI {model}".strip()

    pieces = [type_label]
    if position:
        pieces.append(position)
    pieces.append("DJI")
    if model:
        pieces.append(model)
    return " ".join(pieces).strip()


def category_to_ncm(category: str, type_code: str, model: str) -> tuple[str, str]:
    if category == "DJI > Drones Completos":
        return "8806.91.00", "REVISAR"
    if type_code in {"HELICE"}:
        return "8807.10.00", "OK"
    if type_code in {"BRACO", "SHELL"}:
        return "8807.30.00", "REVISAR"
    if type_code == "HUB":
        return "8504.40.10", "OK"
    if type_code == "CABO-PTZ":
        return "8544.42.00", "OK"
    if type_code == "CONTROLE":
        return "8526.92.00", "OK"
    if type_code == "GPS-IMU":
        return "8526.91.00", "OK"
    return "", "REVISAR"


def apply_ncm(
    category: str, type_code: str, model: str, ncm_index: dict[str, dict]
) -> tuple[str, str, str, str]:
    code, status = category_to_ncm(category, type_code, model)
    if not code:
        return "", "REVISAR", "", ""
    ncm = ncm_index.get(code) or ncm_index.get(re.sub(r"\D", "", code))
    if not ncm:
        return "", "REVISAR", "", ""
    return (
        ncm["codigo"],
        status,
        "Receita Federal/Classif JSON vigente",
        ncm["descricao"],
    )


def build_products(
    raw_products: list[dict], category_map: dict[str, int | None], ncm_index: dict[str, dict]
) -> list[dict]:
    products: list[dict] = []
    used_skus: Counter[str] = Counter()

    for item in raw_products:
        original = item["nome_original"]
        model, model_alerts = detect_model(original)
        type_label, type_code, position, type_alerts = infer_type_position(original)
        category = infer_category(original)
        category_id = category_map.get(category)
        name = build_name(type_label, position, model, original)
        model_sku = MODEL_SKU.get(model, sku_part(model, "SEM-MODELO"))
        position_sku = sku_part(position, "GERAL")
        base_sku = f"BDP-{model_sku}-{type_code}-{position_sku}-L{int(item['linha_ods']):03d}"
        used_skus[base_sku] += 1
        sku = base_sku

        quantity = item["quantidade"]
        cost = item["custo_unitario"]
        price = item["preco_venda"]
        sale_value = money((quantity or Decimal("0")) * (price or Decimal("0")))
        margin = None
        if price and price != 0 and cost is not None:
            margin = ((price - cost) / price * Decimal("100")).quantize(Decimal("0.1"))

        ncm, ncm_status, ncm_source, ncm_description = apply_ncm(
            category, type_code, model, ncm_index
        )

        alerts = [*model_alerts, *type_alerts]
        if not category_id:
            alerts.append("categoria sem ID no mapa Bling")
        if category == "DJI > Outros a Classificar":
            alerts.append("categoria não detectada")
        if price is None:
            alerts.append("sem preço")
        if quantity is None:
            alerts.append("sem estoque")
        elif quantity == 0:
            alerts.append("estoque zero")
        elif quantity < 0:
            alerts.append("quantidade negativa")
        if margin is not None and margin < 0:
            alerts.append("margem negativa")
        elif margin is not None and margin < Decimal("20"):
            alerts.append("margem baixa")
        if not ncm or ncm_status == "REVISAR":
            alerts.append("NCM pendente de validação fiscal")
        alerts.append("GTIN pendente")
        alerts.append("logística pendente: peso/dimensões não informados")
        alerts.append("imagem pendente")
        alerts.append("unidade pendente no ODS")

        bloqueios = []
        if not name:
            bloqueios.append("sem nome_bling")
        if not sku:
            bloqueios.append("sem sku")
        if price is None:
            bloqueios.append("sem preço_venda")
        if not category_id:
            bloqueios.append("sem categoria_bling_id")
        if quantity is not None and quantity < 0:
            bloqueios.append("quantidade negativa")

        if bloqueios:
            status = "BLOQUEADO_REVISAR"
        elif alerts:
            status = "IMPORTAR_COM_ALERTA"
        else:
            status = "PRONTO_PARA_IMPORTAR"

        description = (
            f"Peça/componente DJI para reposição. Compatível com {model}. "
            "Confira modelo e posição da peça antes da compra."
            if model
            else "Peça/componente DJI para reposição. Confira compatibilidade, modelo e posição da peça antes da compra."
        )

        products.append(
            {
                "linha_ods": item["linha_ods"],
                "status_cadastro": status,
                "alertas": "; ".join(dict.fromkeys(alerts)),
                "nome_original": original,
                "nome_bling": name,
                "sku": sku,
                "descricao_curta": description,
                "marca": "DJI",
                "modelo_detectado": model,
                "tipo_peca": type_label,
                "categoria_path": category,
                "categoria_bling_id": category_id,
                "quantidade": quantity,
                "custo_unitario": cost,
                "preco_venda": price,
                "margem_pct": margin,
                "valor_estoque_venda": sale_value,
                "unidade": "",
                "formato_produto": "S",
                "situacao_sugerida": "A" if status != "BLOQUEADO_REVISAR" else "I",
                "peso_liquido": "",
                "peso_bruto": "",
                "largura": "",
                "altura": "",
                "profundidade": "",
                "gtin": "",
                "ncm": ncm,
                "ncm_status": ncm_status,
                "ncm_fonte": ncm_source,
                "ncm_descricao_oficial": ncm_description,
                "url_imagem": "",
                "bling_id": "",
                "status_importacao": "NAO_EXECUTADO",
                "erro_importacao": "",
            }
        )

    sku_counts = Counter(p["sku"] for p in products)
    if any(count > 1 for count in sku_counts.values()):
        suffix_counts: Counter[str] = Counter()
        for product in products:
            if sku_counts[product["sku"]] <= 1:
                continue
            suffix_counts[product["sku"]] += 1
            suffix = chr(ord("A") + suffix_counts[product["sku"]] - 1)
            product["sku"] = f"{product['sku']}-{suffix}"
            product["alertas"] += "; SKU duplicado resolvido com sufixo"

    return products


def validate_invariants(raw_products: list[dict]) -> dict:
    total_qty = sum((p["quantidade"] or Decimal("0")) for p in raw_products)
    total_cost = money(
        sum(
            (p["quantidade"] or Decimal("0")) * (p["custo_unitario"] or Decimal("0"))
            for p in raw_products
        )
    )
    total_sale = money(
        sum(
            (p["quantidade"] or Decimal("0")) * (p["preco_venda"] or Decimal("0"))
            for p in raw_products
        )
    )
    errors = []
    if len(raw_products) != EXPECTED_PRODUCT_COUNT:
        errors.append(f"total de produtos divergente: {len(raw_products)}")
    if total_qty != EXPECTED_TOTAL_QTY:
        errors.append(f"total de unidades divergente: {total_qty}")
    if abs(total_cost - EXPECTED_TOTAL_COST) > MONEY_TOLERANCE:
        errors.append(f"custo total divergente: {total_cost}")
    if abs(total_sale - EXPECTED_TOTAL_SALE) > MONEY_TOLERANCE:
        errors.append(f"venda potencial divergente: {total_sale}")
    return {
        "total_produtos": len(raw_products),
        "total_unidades": total_qty,
        "custo_total": total_cost,
        "venda_potencial": total_sale,
        "errors": errors,
    }


def serializable(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def write_json(path: Path, value) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, default=serializable) + "\n",
        "utf-8",
    )


def add_sheet(wb: Workbook, title: str, rows: list[dict], columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([serializable(row.get(col, "")) for col in columns])
    format_sheet(ws)


def format_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 60)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)


def build_workbooks(
    products: list[dict],
    category_map: dict[str, int | None],
    ncm_rows: list[dict],
    ncm_downloaded_at: str,
    invariants: dict,
) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)

    pendencies = [
        {
            "linha_ods": p["linha_ods"],
            "sku": p["sku"],
            "nome_bling": p["nome_bling"],
            "status_cadastro": p["status_cadastro"],
            "alertas": p["alertas"],
        }
        for p in products
        if p["status_cadastro"] != "PRONTO_PARA_IMPORTAR" or p["alertas"]
    ]
    categories = [
        {
            "categoria_path": path,
            "categoria_bling_id": category_map.get(path),
            "quantidade_produtos": sum(1 for p in products if p["categoria_path"] == path),
            "quantidade_esperada": EXPECTED_DISTRIBUTION.get(path, 0),
            "status": "OK"
            if sum(1 for p in products if p["categoria_path"] == path)
            == EXPECTED_DISTRIBUTION.get(path, 0)
            else "ALERTA_DISTRIBUICAO",
        }
        for path in EXPECTED_DISTRIBUTION
    ]
    status_counts = Counter(p["status_cadastro"] for p in products)
    ncm_counts = Counter(p["ncm_status"] for p in products)
    category_counts = Counter(p["categoria_path"] for p in products)
    resumo = [
        {"indicador": "produtos_lidos", "valor": invariants["total_produtos"]},
        {"indicador": "estoque_total_unidades", "valor": serializable(invariants["total_unidades"])},
        {"indicador": "custo_total", "valor": serializable(invariants["custo_total"])},
        {"indicador": "valor_venda_potencial", "valor": serializable(invariants["venda_potencial"])},
        {"indicador": "ncm_baixado_em", "valor": ncm_downloaded_at},
        {"indicador": "skus_unicos", "valor": len({p["sku"] for p in products}) == len(products)},
    ]
    resumo.extend({"indicador": f"status_{k}", "valor": v} for k, v in status_counts.items())
    resumo.extend({"indicador": f"ncm_status_{k}", "valor": v} for k, v in ncm_counts.items())
    resumo.extend({"indicador": f"categoria_{k}", "valor": v} for k, v in category_counts.items())

    payload_preview = [
        {
            "linha_ods": p["linha_ods"],
            "sku": p["sku"],
            "status_cadastro": p["status_cadastro"],
            "payload_previsto": json.dumps(build_payload_preview(p), ensure_ascii=False),
        }
        for p in products
        if p["status_cadastro"] != "BLOQUEADO_REVISAR"
    ]

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Produtos_Bling", products, MAIN_COLUMNS)
    add_sheet(wb, "Pendencias", pendencies, ["linha_ods", "sku", "nome_bling", "status_cadastro", "alertas"])
    add_sheet(wb, "Categorias", categories, ["categoria_path", "categoria_bling_id", "quantidade_produtos", "quantidade_esperada", "status"])
    add_sheet(wb, "NCM_Auditoria", ncm_rows, ["ncm", "ncm_status", "ncm_fonte", "ncm_descricao_oficial", "produtos", "baixado_em"])
    add_sheet(wb, "Resumo", resumo, ["indicador", "valor"])
    add_sheet(wb, "Payload_Dry_Run", payload_preview, ["linha_ods", "sku", "status_cadastro", "payload_previsto"])
    wb.save(OUTPUT_DIR / "01_produtos_bling_revisao.xlsx")

    wb_import = Workbook()
    wb_import.remove(wb_import.active)
    importable = [p for p in products if p["status_cadastro"] != "BLOQUEADO_REVISAR"]
    add_sheet(wb_import, "Produtos_Bling", importable, MAIN_COLUMNS)
    wb_import.save(OUTPUT_DIR / "02_produtos_bling_importacao.xlsx")

    for filename, rows, columns in [
        ("03_pendencias.xlsx", pendencies, ["linha_ods", "sku", "nome_bling", "status_cadastro", "alertas"]),
        ("04_categorias_usadas.xlsx", categories, ["categoria_path", "categoria_bling_id", "quantidade_produtos", "quantidade_esperada", "status"]),
        ("05_ncm_auditoria.xlsx", ncm_rows, ["ncm", "ncm_status", "ncm_fonte", "ncm_descricao_oficial", "produtos", "baixado_em"]),
    ]:
        single = Workbook()
        single.remove(single.active)
        add_sheet(single, Path(filename).stem, rows, columns)
        single.save(OUTPUT_DIR / filename)

    write_json(OUTPUT_DIR / "produtos_bling_revisao.json", products)
    write_json(OUTPUT_DIR / "category-map.json", category_map)
    write_json(OUTPUT_DIR / "06_payloads_dry_run.json", {"status": "pending_import_products_dry_run", "payloads": []})
    write_json(OUTPUT_DIR / "07_resultado_importacao.json", {"status": "not_executed", "results": []})
    write_report(products, categories, ncm_rows, invariants, ncm_downloaded_at)


def build_payload_preview(product: dict) -> dict:
    payload = {
        "codigo": product["sku"],
        "nome": product["nome_bling"],
        "tipo": "P",
        "situacao": product["situacao_sugerida"],
        "preco": serializable(product["preco_venda"]),
        "marca": "DJI",
        "formato": "S",
        "categoria": {"id": product["categoria_bling_id"]},
        "descricaoCurta": product["descricao_curta"],
    }
    if product["custo_unitario"] is not None:
        payload["precoCusto"] = serializable(product["custo_unitario"])
    if product["ncm"]:
        payload["ncm"] = product["ncm"]
    if product["unidade"]:
        payload["unidade"] = product["unidade"]
    return payload


def write_report(
    products: list[dict],
    categories: list[dict],
    ncm_rows: list[dict],
    invariants: dict,
    ncm_downloaded_at: str,
) -> None:
    status_counts = Counter(p["status_cadastro"] for p in products)
    ncm_counts = Counter(p["ncm_status"] for p in products)
    blocked = [p for p in products if p["status_cadastro"] == "BLOQUEADO_REVISAR"]
    ncm_pending = [p for p in products if p["ncm_status"] == "REVISAR" or not p["ncm"]]
    lines = [
        "# Relatório Final - Pré-Importação Bling Brasil Drones",
        "",
        f"Gerado em: {now_iso()}",
        f"NCM/Classif baixado em: {ncm_downloaded_at}",
        "",
        "## Totais",
        f"- Total de produtos lidos: {invariants['total_produtos']}",
        f"- Total de unidades: {invariants['total_unidades']}",
        f"- Custo total: R$ {invariants['custo_total']}",
        f"- Valor potencial de venda: R$ {invariants['venda_potencial']}",
        f"- Confere 78 produtos: {'sim' if invariants['total_produtos'] == EXPECTED_PRODUCT_COUNT else 'não'}",
        f"- Confere 241 unidades: {'sim' if invariants['total_unidades'] == EXPECTED_TOTAL_QTY else 'não'}",
        f"- SKUs únicos: {'sim' if len({p['sku'] for p in products}) == len(products) else 'não'}",
        "",
        "## Total por status_cadastro",
    ]
    lines.extend(f"- {key}: {value}" for key, value in sorted(status_counts.items()))
    lines.extend(["", "## Total por ncm_status"])
    lines.extend(f"- {key}: {value}" for key, value in sorted(ncm_counts.items()))
    lines.extend(["", "## Total por categoria"])
    lines.extend(
        f"- {row['categoria_path']}: {row['quantidade_produtos']} (esperado {row['quantidade_esperada']})"
        for row in categories
    )
    lines.extend(["", "## NCMs usados"])
    used_ncms = [row for row in ncm_rows if row["ncm"]]
    lines.extend(
        f"- {row['ncm']} ({row['ncm_status']}): {row['produtos']} produtos - {row['ncm_descricao_oficial']}"
        for row in used_ncms
    )
    lines.extend(["", "## Itens com NCM pendente"])
    lines.extend(f"- L{p['linha_ods']} {p['sku']} - {p['nome_bling']}" for p in ncm_pending)
    lines.extend(["", "## Itens bloqueados"])
    if blocked:
        lines.extend(
            f"- L{p['linha_ods']} {p['sku']} - {p['nome_bling']} - {p['alertas']}"
            for p in blocked
        )
    else:
        lines.append("- Nenhum")
    lines.extend(
        [
            "",
            "## Importação",
            "- Itens importados: não executado",
            "- Produtos já existentes: não verificado ainda",
            "- Erros da API: não executado",
        ]
    )
    (OUTPUT_DIR / "08_relatorio_final.md").write_text("\n".join(lines) + "\n", "utf-8")


def build_ncm_audit(products: list[dict], downloaded_at: str) -> list[dict]:
    grouped: dict[tuple[str, str, str, str], list[str]] = {}
    for product in products:
        key = (
            product["ncm"],
            product["ncm_status"],
            product["ncm_fonte"],
            product["ncm_descricao_oficial"],
        )
        grouped.setdefault(key, []).append(product["sku"])
    return [
        {
            "ncm": key[0],
            "ncm_status": key[1],
            "ncm_fonte": key[2],
            "ncm_descricao_oficial": key[3],
            "produtos": len(skus),
            "baixado_em": downloaded_at,
        }
        for key, skus in sorted(grouped.items())
    ]


def main() -> None:
    ods_path = resolve_ods_path()
    raw_products = read_ods_rows(ods_path)
    invariants = validate_invariants(raw_products)
    OUTPUT_DIR.mkdir(exist_ok=True)
    if invariants["errors"]:
        write_json(OUTPUT_DIR / "07_resultado_importacao.json", {"status": "error", "errors": invariants["errors"]})
        raise SystemExit("; ".join(invariants["errors"]))

    ncm_index, downloaded_at = download_ncm_table()
    category_map = load_category_map(resolve_category_map_path())
    products = build_products(raw_products, category_map, ncm_index)
    ncm_audit = build_ncm_audit(products, downloaded_at)
    build_workbooks(products, category_map, ncm_audit, downloaded_at, invariants)
    print(f"Planilhas geradas em {OUTPUT_DIR}")
    print(f"ODS de origem: {ods_path}")
    print(f"Produtos: {len(products)} | Unidades: {invariants['total_unidades']} | Venda: {invariants['venda_potencial']}")


if __name__ == "__main__":
    main()
