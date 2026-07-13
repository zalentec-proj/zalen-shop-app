import json
import re
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "saida_bling"
ODS_FILE = Path("/Users/jeffersonweiberpalombo/Downloads/CADASTPRODUTOS (1).ods")
REVIEW_XLSX = OUT / "01_produtos_bling_revisao.xlsx"
UPDATED_XLSX = OUT / "24_produtos_bling_nomes_originais.xlsx"
AUDIT_XLSX = OUT / "24_auditoria_nomes_originais.xlsx"
DRY_RUN_JSON = OUT / "24_nomes_originais_dry_run.json"

NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


def main():
    assert_file(ODS_FILE)
    assert_file(REVIEW_XLSX)
    OUT.mkdir(parents=True, exist_ok=True)

    ods_products = read_ods_products(ODS_FILE)
    if len(ods_products) != 78:
        raise ValueError(f"ODS invalido: esperado 78 produtos, recebido {len(ods_products)}")

    workbook = openpyxl.load_workbook(REVIEW_XLSX)
    products_sheet = workbook["Produtos_Bling"]
    headers = read_headers(products_sheet)

    required_headers = ["linha_ods", "nome_original", "nome_bling", "sku", "bling_id"]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"Planilha sem colunas obrigatorias: {', '.join(missing)}")

    previous_header = "nome_padronizado_anterior"
    if previous_header not in headers:
        new_column = products_sheet.max_column + 1
        products_sheet.cell(row=1, column=new_column).value = previous_header
        source = products_sheet.cell(row=1, column=headers["nome_bling"])
        target = products_sheet.cell(row=1, column=new_column)
        copy_style(source, target)
        headers = read_headers(products_sheet)

    audit_rows = []
    dry_payloads = []
    products_seen = 0

    for row_idx in range(2, products_sheet.max_row + 1):
        line = products_sheet.cell(row=row_idx, column=headers["linha_ods"]).value
        if line is None:
            continue
        line = int(line)
        ods_product = ods_products.get(line)
        if not ods_product:
            raise ValueError(f"Linha {line} nao encontrada na ODS correta")

        products_seen += 1
        original_ods = ods_product["PRODUTO"]
        original_clean = normalize_visible_name(original_ods)
        current_original = products_sheet.cell(row=row_idx, column=headers["nome_original"]).value
        current_name = products_sheet.cell(row=row_idx, column=headers["nome_bling"]).value
        sku = products_sheet.cell(row=row_idx, column=headers["sku"]).value
        bling_id = products_sheet.cell(row=row_idx, column=headers["bling_id"]).value

        if normalize_visible_name(current_original) != original_clean:
            raise ValueError(
                f"nome_original diverge da ODS na linha {line}: "
                f"{current_original!r} vs {original_ods!r}"
            )

        products_sheet.cell(row=row_idx, column=headers[previous_header]).value = current_name
        products_sheet.cell(row=row_idx, column=headers["nome_bling"]).value = original_clean

        action = "ATUALIZAR_NOME" if bling_id else "SEM_BLING_ID"
        changed = normalize_visible_name(current_name) != original_clean
        audit = {
            "linha_ods": line,
            "sku": sku,
            "bling_id": bling_id,
            "nome_original_ods": original_clean,
            "nome_atual_bling_planilha": current_name,
            "nome_padronizado_anterior": current_name,
            "nome_novo": original_clean,
            "alterado": changed,
            "acao": action if changed else "SEM_ALTERACAO",
            "observacao": "Produto sem bling_id; atualizar apenas apos cadastro no Bling" if not bling_id else "",
        }
        audit_rows.append(audit)

        if bling_id and changed:
            dry_payloads.append(
                {
                    "linha_ods": line,
                    "sku": sku,
                    "bling_id": int(bling_id),
                    "nome_atual": current_name,
                    "nome_novo": original_clean,
                    "payload_patch_sugerido": {"nome": original_clean},
                }
            )

    if products_seen != 78:
        raise ValueError(f"Planilha invalida: esperado 78 produtos, recebido {products_seen}")

    workbook.save(UPDATED_XLSX)
    write_audit_workbook(audit_rows)
    dry_run = {
        "dryRun": True,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceOds": str(ODS_FILE),
        "sourceWorkbook": str(REVIEW_XLSX),
        "updatedWorkbook": str(UPDATED_XLSX),
        "auditWorkbook": str(AUDIT_XLSX),
        "rules": {
            "onlyFieldToUpdateInBling": "nome",
            "preserve": [
                "sku",
                "preco",
                "precoCusto",
                "estoque",
                "categoria",
                "ncm",
                "imagens",
                "peso",
                "dimensoes",
                "descricao",
            ],
            "trimOnlyOuterWhitespace": True,
        },
        "summary": {
            "produtosOds": len(ods_products),
            "produtosPlanilha": products_seen,
            "nomesDivergentes": sum(1 for row in audit_rows if row["alterado"]),
            "comBlingIdParaAtualizar": len(dry_payloads),
            "semBlingId": sum(1 for row in audit_rows if not row["bling_id"]),
        },
        "payloads": dry_payloads,
        "audit": audit_rows,
    }
    DRY_RUN_JSON.write_text(json.dumps(dry_run, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(dry_run["summary"], ensure_ascii=False, indent=2))


def read_ods_products(path):
    rows = read_ods_rows(path)
    headers = [normalize_header(value) for value in rows[0]]
    products = {}
    for line, row in enumerate(rows[1:], start=2):
        values = row + [""] * (len(headers) - len(row))
        item = dict(zip(headers, values))
        if not item.get("PRODUTO"):
            continue
        products[line] = item
    return products


def read_ods_rows(path):
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))
    sheet = root.find(".//table:table", NS)
    if sheet is None:
        raise ValueError("ODS sem tabela")

    rows = []
    for row in sheet.findall("table:table-row", NS):
        output = []
        repeated_rows = int(row.attrib.get(f"{{{NS['table']}}}number-rows-repeated", "1"))
        if repeated_rows > 100:
            repeated_rows = 1
        for cell in row.findall("table:table-cell", NS):
            repeated_cols = int(cell.attrib.get(f"{{{NS['table']}}}number-columns-repeated", "1"))
            if repeated_cols > 50:
                repeated_cols = 1
            output.extend([cell_text(cell)] * repeated_cols)
        if any(str(value).strip() for value in output):
            for _ in range(repeated_rows):
                rows.append(output)
    return rows


def cell_text(cell):
    parts = []
    for paragraph in cell.findall(".//text:p", NS):
        parts.append("".join(paragraph.itertext()))
    return "\n".join(parts)


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def normalize_visible_name(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def read_headers(sheet):
    return {str(cell.value): cell.column for cell in sheet[1] if cell.value}


def copy_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def write_audit_workbook(rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Auditoria_Nomes"
    headers = [
        "linha_ods",
        "sku",
        "bling_id",
        "nome_original_ods",
        "nome_atual_bling_planilha",
        "nome_novo",
        "acao",
        "observacao",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)
    workbook.save(AUDIT_XLSX)


def assert_file(path):
    if not path.exists():
        raise FileNotFoundError(str(path))


if __name__ == "__main__":
    main()
